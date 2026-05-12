"""
add_country_summaries.py

Reads countries.xlsx, fetches each country/territory Wikipedia intro,
asks Claude to write a concise summary, and writes it into column E.

Usage:
    pip install anthropic openpyxl requests
    export ANTHROPIC_API_KEY="sk-ant-..."
    python add_country_summaries.py

Notes:
- Resume-safe: rows that already have a summary in column E are skipped.
- Auto-save: saves every 10 successful rows.
"""

from __future__ import annotations

import os
import time
from typing import Dict

import anthropic
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_FILE = "countries.xlsx"
SHEET_NAME = "All Countries"
COUNTRY_COL = 2          # Column B: country name
SUMMARY_COL = 5          # Column E: output summary
HEADER_ROW = 1
SAVE_EVERY = 10
SLEEP_BETWEEN = 0.3      # seconds between model calls

WIKI_API = "https://en.wikipedia.org/api/rest_v1/page/summary/{}"
MODEL = "claude-3-5-haiku-latest"

SYSTEM_PROMPT = (
    "You are a concise travel and geography writer. "
    "Write a 2-3 sentence summary of the given country or territory: "
    "what it is known for, where it is, and one distinctive fact. "
    "Be factual, neutral, and engaging. Do not start with the country name."
)

# Input-name normalizations to improve Wikipedia hits.
WIKI_ALIASES: Dict[str, str] = {
    "Türkiye (Turkey)": "Turkey",
    "Czechia (Czech Republic)": "Czech Republic",
    "Ivory Coast (Côte d'Ivoire)": "Ivory Coast",
    "Vatican City (Holy See)": "Vatican City",
    "Palestine (UN observer state)": "State of Palestine",
    "Russia (spans Europe & Asia)": "Russia",
}


def normalize_for_wikipedia(name: str) -> str:
    return WIKI_ALIASES.get(name, name)


def fetch_wikipedia_intro(country: str) -> str:
    """Return the Wikipedia intro extract for a country/territory."""
    name = normalize_for_wikipedia(country)

    def _fetch(query: str) -> str:
        url = WIKI_API.format(requests.utils.quote(query))
        response = requests.get(
            url,
            timeout=10,
            headers={"User-Agent": "CountrySummaryBot/1.0"},
        )
        if response.status_code != 200:
            return ""
        payload = response.json()
        return payload.get("extract", "") or ""

    try:
        text = _fetch(name)
        if text:
            return text
        # Retry with "country" suffix for ambiguous pages.
        return _fetch(f"{name} country")
    except Exception as exc:  # keep run resilient row-by-row
        print(f"    Wikipedia fetch error for '{country}': {exc}")
        return ""


def ask_claude(client: anthropic.Anthropic, country: str, wiki_text: str) -> str:
    """Request a short summary from Claude."""
    if wiki_text:
        user_prompt = (
            f"Country/territory: {country}\n\n"
            f"Wikipedia excerpt:\n{wiki_text[:3000]}\n\n"
            "Write a 2-3 sentence summary."
        )
    else:
        user_prompt = (
            f"Country/territory: {country}\n\n"
            "No Wikipedia text was available. "
            "Write a 2-3 sentence summary from your own knowledge."
        )

    response = client.messages.create(
        model=MODEL,
        max_tokens=200,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_prompt}],
    )
    return response.content[0].text.strip()


def main() -> None:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise SystemExit("ERROR: Set ANTHROPIC_API_KEY before running.")

    if not os.path.exists(EXCEL_FILE):
        raise SystemExit(f"ERROR: '{EXCEL_FILE}' not found in current directory.")

    client = anthropic.Anthropic(api_key=api_key)
    workbook = load_workbook(EXCEL_FILE)
    worksheet = workbook[SHEET_NAME]

    # Header formatting for Summary column.
    header_cell = worksheet.cell(row=HEADER_ROW, column=SUMMARY_COL)
    if not header_cell.value:
        header_cell.value = "Summary"
        thin = Side(style="thin", color="CCCCCC")
        header_cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill("solid", start_color="2E4057")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    worksheet.column_dimensions["E"].width = 80

    max_row = worksheet.max_row
    processed = 0
    skipped = 0

    print(f"Found {max_row - HEADER_ROW} rows. Starting...\n")

    for row in range(HEADER_ROW + 1, max_row + 1):
        country_cell = worksheet.cell(row=row, column=COUNTRY_COL)
        summary_cell = worksheet.cell(row=row, column=SUMMARY_COL)
        country = country_cell.value

        if not country:
            continue

        if summary_cell.value and str(summary_cell.value).strip():
            skipped += 1
            continue

        print(f"[{row - HEADER_ROW}/{max_row - HEADER_ROW}] {country}")
        wiki_text = fetch_wikipedia_intro(str(country))
        print(f"    {'✓' if wiki_text else '✗'} Wikipedia")

        try:
            summary = ask_claude(client, str(country), wiki_text)
            summary_cell.value = summary
            summary_cell.alignment = Alignment(wrap_text=True, vertical="top")
            worksheet.row_dimensions[row].height = 60
            processed += 1
            print(f"    → {summary[:90]}{'…' if len(summary) > 90 else ''}")
        except Exception as exc:
            summary_cell.value = f"ERROR: {exc}"
            print(f"    Claude error: {exc}")

        if processed > 0 and processed % SAVE_EVERY == 0:
            workbook.save(EXCEL_FILE)
            print(f"    💾 Saved after {processed} summaries.\n")

        time.sleep(SLEEP_BETWEEN)

    workbook.save(EXCEL_FILE)
    print(f"\nDone! Wrote {processed} summaries, skipped {skipped} prefilled rows.")
    print(f"Saved to: {os.path.abspath(EXCEL_FILE)}")


if __name__ == "__main__":
    main()
